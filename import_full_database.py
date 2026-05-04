"""
Import Full Vigilent Database into QGIS2Web Map
=================================================
Reads the complete database Excel, geocodes missing coordinates,
and generates updated GeoJSON data files for all regions.

Usage:
    python3 import_full_database.py
"""

import json
import re
import os
import openpyxl
import urllib.request
import urllib.parse
import time

EXCEL_PATH = "/Users/adamtang/Downloads/VIGILENT Data Center Database Backup.xlsx"
DATA_DIR = "qgis2web_2026_04_01-14_53_08_869925/data"

# Region → GeoJSON filename mapping
REGION_FILES = {
    "US": "VigilentDataCenterDatabaseUS_22.js",
    "Canada": "VigilentDataCenterDatabaseCanada_21.js",
    "Europe": "VigilentDataCenterDatabaseEurope_20.js",
    "Brazil": "VigilentDataCenterDatabaseBrazil_19.js",
}

# Map "Other" sheet countries to existing regions or new ones
COUNTRY_TO_REGION = {
    "Brazil": "Brazil",
    "India": "Other",
    "Singapore": "Other",
    "Thailand": "Other",
}

# Known city coordinates for geocoding fallback
CITY_COORDS = {
    # Europe - major DC hub cities
    ("London", "UK"): (51.5074, -0.1278),
    ("London", "N/A"): (51.5074, -0.1278),
    ("London", "London"): (51.5074, -0.1278),
    ("Manchester", "North West"): (53.4808, -2.2426),
    ("Manchester", "Greater Manchester"): (53.4808, -2.2426),
    ("Amsterdam", None): (52.3676, 4.9041),
    ("Amsterdam", "N/A"): (52.3676, 4.9041),
    ("Amsterdam", "North Holland"): (52.3676, 4.9041),
    ("Frankfurt", None): (50.1109, 8.6821),
    ("Frankfurt", "Hesse"): (50.1109, 8.6821),
    ("Frankfurt", "N/A"): (50.1109, 8.6821),
    ("Paris", None): (48.8566, 2.3522),
    ("Paris", "Ile-de-France"): (48.8566, 2.3522),
    ("Paris", "N/A"): (48.8566, 2.3522),
    ("Dublin", None): (53.3498, -6.2603),
    ("Dublin", "N/A"): (53.3498, -6.2603),
    ("Dublin", "Leinster"): (53.3498, -6.2603),
    ("Stockholm", None): (59.3293, 18.0686),
    ("Stockholm", "N/A"): (59.3293, 18.0686),
    ("Stockholm", "Stockholm"): (59.3293, 18.0686),
    ("Madrid", None): (40.4168, -3.7038),
    ("Madrid", "Madrid"): (40.4168, -3.7038),
    ("Madrid", "N/A"): (40.4168, -3.7038),
    ("Copenhagen", None): (55.6761, 12.5683),
    ("Copenhagen", "N/A"): (55.6761, 12.5683),
    ("Copenhagen", "Capital Region"): (55.6761, 12.5683),
    ("Corsham", "South West"): (51.4311, -2.1877),
    ("Newport", "South Wales"): (51.5842, -2.9977),
    ("Slough", None): (51.5105, -0.5950),
    ("Slough", "Berkshire"): (51.5105, -0.5950),
    # India
    ("Mumbai", "Maharashtra"): (19.0760, 72.8777),
    ("Mumbai", None): (19.0760, 72.8777),
    ("Chennai", "Tamil Nadu"): (13.0827, 80.2707),
    ("Chennai", None): (13.0827, 80.2707),
    ("Pune", "Maharashtra"): (18.5204, 73.8567),
    ("Pune", None): (18.5204, 73.8567),
    ("Hyderabad", "Telangana"): (17.3850, 78.4867),
    ("Noida", "Uttar Pradesh"): (28.5355, 77.3910),
    ("Bangalore", "Karnataka"): (12.9716, 77.5946),
    ("Navi Mumbai", "Maharashtra"): (19.0330, 73.0297),
    ("Delhi", None): (28.6139, 77.2090),
    ("Kolkata", None): (22.5726, 88.3639),
    # Singapore
    ("Singapore", "N/A"): (1.3521, 103.8198),
    ("Singapore", None): (1.3521, 103.8198),
    # Europe - Ireland
    ("Kilmahuddrick", ""): (53.3200, -6.3900),
    ("Kilcarbery", ""): (53.3080, -6.4150),
    ("Clonshaugh", ""): (53.4050, -6.2000),
    ("Ballybane", ""): (53.2750, -8.9850),
    ("Ballycoolin", ""): (53.3950, -6.3550),
    # Europe - Denmark
    ("Ballerup", ""): (55.7317, 12.3637),
    # Europe - UK
    ("Watford", ""): (51.6565, -0.3903),
    # Europe - France
    ("La Courneuve", ""): (48.9279, 2.3962),
    ("Les Ulis", ""): (48.6778, 2.1700),
    # Europe - Netherlands
    ("Haarlem", ""): (52.3874, 4.6462),
    ("Schiphol-Rijk", ""): (52.2950, 4.7650),
    # Europe - Sweden
    ("Kista", ""): (59.4030, 17.9440),
    ("Sollentuna", ""): (59.4280, 17.9510),
    ("Skarholmen", ""): (59.2770, 17.9070),
    ("Vasby", ""): (59.5184, 17.9137),
    ("Skondal", ""): (59.2580, 18.1290),
    ("Conapto Stockholm 1 City", ""): (59.3293, 18.0686),
    # Europe - Ireland (None state)
    ("Dublin", None): (53.3498, -6.2603),
    ("Dublin", "Leinster"): (53.3498, -6.2603),
    ("Dublin", ""): (53.3498, -6.2603),
    ("Stockholm", ""): (59.3293, 18.0686),
    ("Stockholm ", None): (59.3293, 18.0686),
    ("Stockholm ", ""): (59.3293, 18.0686),
    # Facility code inference (no city/country in spreadsheet)
    ("", ""): None,  # skip truly empty
    # India
    ("Navi Mumbai", None): (19.0330, 73.0297),
    ("Navi Mumbai", "Maharashtra"): (19.0330, 73.0297),
    # Thailand
    ("Bangkok", "N/A"): (13.7563, 100.5018),
    ("Bangkok", None): (13.7563, 100.5018),
    # Brazil
    ("São Paulo", "São Paulo"): (-23.5505, -46.6333),
    ("Rio de Janeiro", "Rio de Janeiro"): (-22.9068, -43.1729),
    ("Vinhedo", "Rio de Janeiro"): (-23.0297, -46.9753),
    ("Palmas", "Tocantins"): (-10.1689, -48.3317),
    ("Fortaleza", "Ceará"): (-3.7172, -38.5433),
    ("Cotia", "São Paulo"): (-23.6035, -46.9192),
    ("Campinas", "São Paulo"): (-22.9099, -47.0626),
    ("Porto Alegre", "Rio Grande do Sul"): (-30.0346, -51.2177),
}


def geocode_nominatim(city, state, country):
    """Geocode using OpenStreetMap Nominatim (free, no API key)."""
    parts = [p for p in [city, state, country] if p and p != "N/A"]
    query = ", ".join(parts)
    url = "https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode({
        "q": query, "format": "json", "limit": 1
    })
    req = urllib.request.Request(url, headers={"User-Agent": "VigilentMapBuilder/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
            if data:
                return float(data[0]["lat"]), float(data[0]["lon"])
    except Exception as e:
        print(f"    Geocode failed for '{query}': {e}")
    return None, None


def get_coords(dc, country_name):
    """Get coordinates, using fallback geocoding if needed."""
    lat = dc.get("Latitude")
    lng = dc.get("Longitude")

    if lat is not None and lng is not None:
        try:
            return float(lat), float(lng)
        except (ValueError, TypeError):
            pass

    city = str(dc.get("City", "") or "").strip()
    state = str(dc.get("State/Province", "") or "").strip()

    # Try city coords lookup
    for key in [(city, state), (city, country_name), (city, None), (city, "N/A")]:
        if key in CITY_COORDS:
            lat, lng = CITY_COORDS[key]
            # Add small offset to avoid stacking
            import random
            random.seed(hash(dc.get("Name", "")))
            lat += random.uniform(-0.02, 0.02)
            lng += random.uniform(-0.02, 0.02)
            return lat, lng

    # Try Nominatim
    print(f"    Geocoding: {dc.get('Name')} ({city}, {state}, {country_name})")
    lat, lng = geocode_nominatim(city, state, country_name)
    if lat and lng:
        time.sleep(1.1)  # Nominatim rate limit
        return lat, lng

    return None, None


def read_excel():
    """Read all DCs from the Excel database."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    all_dcs = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        headers = [h for h in headers if h]

        dcs = []
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row, 1).value
            if not name or not str(name).strip():
                continue
            dc = {}
            for c, h in enumerate(headers):
                dc[h] = ws.cell(row, c + 1).value
            dcs.append(dc)
        all_dcs[sheet_name] = dcs

    return all_dcs


def build_geojson_feature(dc, region, country_name):
    """Build a GeoJSON feature for a data center."""
    lat, lng = get_coords(dc, country_name)
    if lat is None or lng is None:
        return None

    name = str(dc.get("Name", "")).strip()
    props = {
        "Name": name,
        "Country": str(dc.get("Country", country_name) or country_name).strip(),
        "City": str(dc.get("City", "") or "").strip(),
        "State/Province": str(dc.get("State/Province", "") or "").strip(),
        "Operator": str(dc.get("Operator", "") or "").strip(),
        "Size (sq ft)": dc.get("Size (sq ft)"),
        "Size (MW)": dc.get("Size (MW)"),
        "Latitude": lat,
        "Longitude": lng,
        "Operational Status": str(dc.get("Operational Status", "") or "").strip(),
    }

    return {
        "type": "Feature",
        "properties": props,
        "geometry": {
            "type": "Point",
            "coordinates": [lng, lat]
        }
    }


def load_existing_geojson(filepath):
    """Load existing GeoJSON from a qgis2web .js file."""
    with open(filepath, "r") as f:
        content = f.read()
    match = re.match(r'var\s+(json_\w+)\s*=\s*', content)
    if not match:
        return None, None
    var_name = match.group(1)
    json_str = content[match.end():].rstrip().rstrip(";")
    return var_name, json.loads(json_str)


def write_geojson(filepath, var_name, geojson):
    """Write GeoJSON as qgis2web .js file."""
    json_str = json.dumps(geojson, separators=(",", ":"))
    with open(filepath, "w") as f:
        f.write(f"var {var_name} = {json_str};")


def main():
    print("=" * 60)
    print("  IMPORT FULL DATABASE INTO MAP")
    print("=" * 60)

    db = read_excel()

    # Determine region for each DC
    region_dcs = {"US": [], "Canada": [], "Europe": [], "Brazil": [], "Other": []}

    # Process US, Europe, Canada sheets directly
    for sheet in ["US", "Europe", "Canada"]:
        for dc in db.get(sheet, []):
            region_dcs[sheet].append((dc, sheet if sheet != "Europe" else "Europe"))

    # Process "Other" sheet by country
    for dc in db.get("Other", []):
        country = str(dc.get("Country", "")).strip()
        region = COUNTRY_TO_REGION.get(country, "Other")
        region_dcs[region].append((dc, country))

    # Now build/update GeoJSON for each region
    for region, items in region_dcs.items():
        if not items:
            continue

        fname = REGION_FILES.get(region)
        if fname:
            filepath = os.path.join(DATA_DIR, fname)
            var_name, existing = load_existing_geojson(filepath)
            existing_names = {f["properties"]["Name"] for f in existing["features"]}
        else:
            # New region — skip for now (Other includes India, Singapore, Thailand)
            # These need new layer files which is complex
            print(f"\n  {region}: {len(items)} DCs — SKIPPED (no existing map layer)")
            continue

        print(f"\n  {region}: {len(items)} DCs in database, {len(existing_names)} in map")

        added = 0
        skipped_no_coords = 0
        skipped_existing = 0

        for dc, country_name in items:
            name = str(dc.get("Name", "")).strip()
            if name in existing_names:
                skipped_existing += 1
                continue

            feature = build_geojson_feature(dc, region, country_name)
            if feature is None:
                skipped_no_coords += 1
                print(f"    SKIP (no coords): {name}")
                continue

            existing["features"].append(feature)
            existing_names.add(name)
            added += 1

        write_geojson(filepath, var_name, existing)
        new_size = os.path.getsize(filepath) / 1024
        print(f"    Added: {added}, Already existed: {skipped_existing}, "
              f"No coords: {skipped_no_coords}")
        print(f"    File: {fname} ({new_size:.0f} KB)")

    # Summary
    print("\n" + "=" * 60)
    total = sum(len(items) for items in region_dcs.values())
    print(f"  Total DCs processed: {total}")
    print("  Done!")


if __name__ == "__main__":
    main()
