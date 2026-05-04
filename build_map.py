"""
Build Map Data — Dynamic Pipeline
====================================
Reads the full Vigilent database (Excel), runs composite scoring on ALL
data centers (US + international), and generates GeoJSON data files for
the QGIS2Web map.

Usage:
    python3 build_map.py

When new data centers are added to the Excel file, re-run this script
to regenerate all map data with scores.
"""

import json
import os
import re
import random
import openpyxl
from vigilent_engine import compute_score, SCORING_CONFIG
from operator_tiers import tier_for_operator, opex_pct_for_operator

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════════════════════

EXCEL_PATH = "/Users/adamtang/Downloads/VIGILENT Data Center Database Backup.xlsx"
DATA_DIR = "qgis2web_2026_04_01-14_53_08_869925/data"

VIGILENT_PARAMS = {
    "investment_cost": 1_500_000,
    "energy_reduction_pct": 0.10,
    "water_reduction_pct": 0.05,
    "num_years": 1,
}

DEFAULTS = {
    "baseline_pue": 1.55,
    "load_growth_rate": 0.10,
    "energy_pct_opex": 0.40,    # wholesale-colo midline; operator-tier override applied per-DC
    "capacity_factor": 0.70,    # Uptime Institute / IDC avg IT load utilization
}

# Region electricity rate property keys (from choropleth joins in GeoJSON)
REGION_ELEC_KEYS = {
    "US": "Electricity By State_Commercial Electricity Rate Per State (¢/kWh)",
    "Europe": "Europe Statistics_Commercial Electricity Rate (¢/kWh)",
    "Canada": "Canada Statistics_Commercial Electricity Rate (¢/kWh)",
    "Brazil": "Brazil Commercial Electricity Rates (¢/kWh)_Brazil Statistics_Commercial Electricity Rate (¢/kWh)",
    "Other": None,
}
REGION_WATER_KEYS = {
    "US": "Water Cost By State_Commercial Water Price ($/1000 gallons)",
    "Europe": "Europe Statistics_Commercial Water Rate ($/1000 gallons)",
    "Canada": "Canada Statistics_Commercial Water Rate ($/1000 gallons)",
    "Brazil": "Brazil Commercial Electricity Rates (¢/kWh)_Brazil Statistics_Commercial Water Rate ($/1000 gallons)",
    "Other": None,
}
REGION_REG_KEYS = {
    "US": "Regulations_Regulations:",
    "Europe": "Europe Statistics_Regulations",
    "Canada": "Canada Statistics_Regulations",
    "Brazil": "Brazil Commercial Electricity Rates (¢/kWh)_Brazil Statistics_Regulations",
    "Other": None,
}

# State-level electricity rates for US scoring ($/kWh)
US_STATE_ELEC = {
    "TX": 0.0912, "NY": 0.2254, "CO": 0.1332, "NV": 0.0991, "OH": 0.1155,
    "MA": 0.2340, "CA": 0.2500, "VA": 0.0973, "WA": 0.1100, "OR": 0.1136,
    "MN": 0.1322, "MS": 0.1267, "NJ": 0.1600, "MI": 0.1267, "FL": 0.1294,
    "GA": 0.1165, "IL": 0.1401, "PA": 0.1212, "NC": 0.1009, "AZ": 0.1295,
    "CT": 0.2389, "MD": 0.1518, "SC": 0.1061, "TN": 0.1200, "IN": 0.1210,
    "WI": 0.1314, "MO": 0.1157, "KY": 0.1064, "AL": 0.1446, "LA": 0.1039,
    "IA": 0.1320, "OK": 0.1004, "KS": 0.1203, "AR": 0.1080, "UT": 0.1001,
    "NE": 0.0958, "WV": 0.1090, "NM": 0.1224, "ND": 0.0744, "SD": 0.1099,
    "MT": 0.1055, "ID": 0.0983, "WY": 0.0979, "HI": 0.3879, "ME": 0.1620,
    "NH": 0.1950, "VT": 0.1933, "RI": 0.2160, "DE": 0.1344, "DC": 0.1400,
    "AK": 0.2100,
}

# Regional average electricity rates for international scoring ($/kWh)
REGION_DEFAULT_ELEC = {
    "Europe": 0.18,
    "Canada": 0.08,
    "Brazil": 0.15,
    "India": 0.08,
    "Singapore": 0.15,
    "Thailand": 0.12,
    "Other": 0.12,
}

# Country → region mapping for "Other" sheet
COUNTRY_TO_REGION = {
    "Brazil": "Brazil",
    "India": "Other",
    "Singapore": "Other",
    "Thailand": "Other",
}

# City coordinate fallbacks for geocoding
CITY_COORDS = {
    # Europe
    "London": (51.5074, -0.1278), "Manchester": (53.4808, -2.2426),
    "Amsterdam": (52.3676, 4.9041), "Frankfurt": (50.1109, 8.6821),
    "Paris": (48.8566, 2.3522), "Dublin": (53.3498, -6.2603),
    "Stockholm": (59.3293, 18.0686), "Madrid": (40.4168, -3.7038),
    "Copenhagen": (55.6761, 12.5683), "Corsham": (51.4311, -2.1877),
    "Newport": (51.5842, -2.9977), "Slough": (51.5105, -0.5950),
    "Watford": (51.6565, -0.3903), "La Courneuve": (48.9279, 2.3962),
    "Les Ulis": (48.6778, 2.1700), "Haarlem": (52.3874, 4.6462),
    "Kista": (59.4030, 17.9440), "Sollentuna": (59.4280, 17.9510),
    "Ballerup": (55.7317, 12.3637),
    "Kilmahuddrick": (53.3200, -6.3900), "Kilcarbery": (53.3080, -6.4150),
    "Clonshaugh": (53.4050, -6.2000), "Ballybane": (53.2750, -8.9850),
    "Ballycoolin": (53.3950, -6.3550),
    "Skarholmen": (59.2770, 17.9070), "Vasby": (59.5184, 17.9137),
    "Skondal": (59.2580, 18.1290), "Schiphol-Rijk": (52.2950, 4.7650),
    # India
    "Mumbai": (19.0760, 72.8777), "Chennai": (13.0827, 80.2707),
    "Pune": (18.5204, 73.8567), "Hyderabad": (17.3850, 78.4867),
    "Noida": (28.5355, 77.3910), "Bangalore": (12.9716, 77.5946),
    "Navi Mumbai": (19.0330, 73.0297), "Delhi": (28.6139, 77.2090),
    "Kolkata": (22.5726, 88.3639),
    # Singapore
    "Singapore": (1.3521, 103.8198),
    # Thailand
    "Bangkok": (13.7563, 100.5018), "Chon Buri": (13.3611, 100.9847),
    "Nonthaburi": (13.8621, 100.5144),
    # Brazil
    "São Paulo": (-23.5505, -46.6333), "Rio de Janeiro": (-22.9068, -43.1729),
    "Vinhedo": (-23.0297, -46.9753), "Palmas": (-10.1689, -48.3317),
    "Fortaleza": (-3.7172, -38.5433), "Cotia": (-23.6035, -46.9192),
    "Campinas": (-22.9099, -47.0626), "Porto Alegre": (-30.0346, -51.2177),
    # Canada
    "Toronto": (43.6532, -79.3832), "Montreal": (45.5017, -73.5673),
    "Calgary": (51.0447, -114.0719), "Vancouver": (49.2827, -123.1207),
    "Quebec City": (46.8139, -71.2080), "Ottawa": (45.4215, -75.6972),
    "Edmonton": (53.5461, -113.4938), "Halifax": (44.6488, -63.5752),
    "Brampton": (43.7315, -79.7624), "Markham": (43.8561, -79.3370),
    "Regina": (50.4452, -104.6189),
}

# GeoJSON output files
REGION_FILES = {
    "US": ("VigilentDataCenterDatabaseUS_22.js", "json_VigilentDataCenterDatabaseUS_22"),
    "Canada": ("VigilentDataCenterDatabaseCanada_21.js", "json_VigilentDataCenterDatabaseCanada_21"),
    "Europe": ("VigilentDataCenterDatabaseEurope_20.js", "json_VigilentDataCenterDatabaseEurope_20"),
    "Brazil": ("VigilentDataCenterDatabaseBrazil_19.js", "json_VigilentDataCenterDatabaseBrazil_19"),
    "Other": ("VigilentDataCenterDatabaseOther_23.js", "json_VigilentDataCenterDatabaseOther_23"),
}


# ═══════════════════════════════════════════════════════════════════════════════
# SCORING
# ═══════════════════════════════════════════════════════════════════════════════

def classify_score(score):
    if score >= 75: return "Excellent"
    elif score >= 50: return "Good"
    elif score >= 25: return "Moderate"
    else: return "Low"


def get_electricity_price(dc, region):
    """Get electricity price in $/kWh for scoring."""
    # US: use state lookup
    if region == "US":
        state = str(dc.get("State/Province", "") or "").strip()
        if state in US_STATE_ELEC:
            return US_STATE_ELEC[state]
        return 0.12  # fallback

    # International: try to get from joined data (¢/kWh → $/kWh)
    elec_key = REGION_ELEC_KEYS.get(region)
    if elec_key and dc.get(elec_key) is not None:
        try:
            return float(dc[elec_key]) / 100.0
        except (ValueError, TypeError):
            pass

    # Fallback to regional average
    country = str(dc.get("Country", "") or "").strip()
    return REGION_DEFAULT_ELEC.get(country, REGION_DEFAULT_ELEC.get(region, 0.12))


def score_dc(dc, region):
    """Run composite scoring on a single data center."""
    mw = dc.get("Size (MW)")
    if mw is None:
        try:
            mw = float(str(dc.get("Size (MW)", "0")).replace(",", "").strip())
        except (ValueError, TypeError):
            return None
    else:
        try:
            mw = float(mw)
        except (ValueError, TypeError):
            return None

    if mw <= 0:
        return None

    elec_price = get_electricity_price(dc, region)
    operator_name = str(dc.get("Operator", "") or "").strip()
    op_tier = tier_for_operator(operator_name)
    energy_pct_opex = opex_pct_for_operator(operator_name)

    try:
        result = compute_score(
            dc_size_mw=mw,
            baseline_pue=DEFAULTS["baseline_pue"],
            electricity_price=elec_price,
            load_growth_rate=DEFAULTS["load_growth_rate"],
            energy_pct_opex=energy_pct_opex,
            capacity_factor=DEFAULTS["capacity_factor"],
            **VIGILENT_PARAMS,
        )
    except Exception as e:
        print(f"    Score error for {dc.get('Name', '?')}: {e}")
        return None

    composite = result["composite_score"]
    return {
        "composite_score": round(composite, 2),
        "classification": classify_score(composite),
        "savings_per_mw": round(result["savings_per_mw"], 2),
        "payback_years": round(result["payback_period_years"], 3),
        "impact_on_opex_pct": round(result["impact_on_opex_pct"] * 100, 2),
        "estimated_savings": round(result["estimated_savings"], 2),
        "electricity_price_used": round(elec_price, 4),
        "operator_tier": op_tier,
        "energy_pct_opex": round(energy_pct_opex, 3),
        "capacity_factor": DEFAULTS["capacity_factor"],
    }


# ═══════════════════════════════════════════════════════════════════════════════
# GEOCODING
# ═══════════════════════════════════════════════════════════════════════════════

def get_coords(dc):
    """Get coordinates with city fallback."""
    lat = dc.get("Latitude")
    lng = dc.get("Longitude")
    if lat is not None and lng is not None:
        try:
            return float(lat), float(lng)
        except (ValueError, TypeError):
            pass

    city = str(dc.get("City", "") or "").strip()
    if city in CITY_COORDS:
        base_lat, base_lng = CITY_COORDS[city]
        random.seed(hash(str(dc.get("Name", ""))))
        return base_lat + random.uniform(-0.02, 0.02), base_lng + random.uniform(-0.02, 0.02)

    return None, None


# ═══════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ═══════════════════════════════════════════════════════════════════════════════

def load_existing_geojson(region):
    """Load existing GeoJSON to preserve joined choropleth data for US DCs."""
    fname, var_name = REGION_FILES[region]
    filepath = os.path.join(DATA_DIR, fname)
    if not os.path.exists(filepath):
        return {}
    with open(filepath) as f:
        content = f.read()
    match = re.match(r'var\s+json_\w+\s*=\s*', content)
    if not match:
        return {}
    geojson = json.loads(content[match.end():].rstrip().rstrip(";"))
    # Build lookup by name
    lookup = {}
    for feat in geojson["features"]:
        name = feat["properties"].get("Name", "")
        if name:
            lookup[name] = feat["properties"]
    return lookup


def read_excel():
    """Read all DCs from Excel, organized by map region."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    region_dcs = {"US": [], "Canada": [], "Europe": [], "Brazil": [], "Other": []}

    for sheet_name in ["US", "Europe", "Canada"]:
        ws = wb[sheet_name]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        headers = [h for h in headers if h]
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row, 1).value
            if not name or not str(name).strip():
                continue
            dc = {}
            for c, h in enumerate(headers):
                dc[h] = ws.cell(row, c + 1).value
            region_dcs[sheet_name].append(dc)

    # "Other" sheet → split by country
    ws = wb["Other"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    headers = [h for h in headers if h]
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 1).value
        if not name or not str(name).strip():
            continue
        dc = {}
        for c, h in enumerate(headers):
            dc[h] = ws.cell(row, c + 1).value
        country = str(dc.get("Country", "") or "").strip()
        region = COUNTRY_TO_REGION.get(country, "Other")
        region_dcs[region].append(dc)

    return region_dcs


# ═══════════════════════════════════════════════════════════════════════════════
# GEOJSON GENERATION
# ═══════════════════════════════════════════════════════════════════════════════

def build_feature(dc, region, score_result, existing_props=None):
    """Build a GeoJSON feature with scoring data."""
    lat, lng = get_coords(dc)
    if lat is None or lng is None:
        return None

    name = str(dc.get("Name", "")).strip()
    props = {
        "Name": name,
        "Country": str(dc.get("Country", "") or "").strip(),
        "City": str(dc.get("City", "") or "").strip(),
        "State/Province": str(dc.get("State/Province", "") or "").strip(),
        "Operator": str(dc.get("Operator", "") or "").strip(),
        "Size (sq ft)": dc.get("Size (sq ft)"),
        "Size (MW)": dc.get("Size (MW)"),
        "Latitude": lat,
        "Longitude": lng,
        "Operational Status": str(dc.get("Operational Status", "") or "").strip(),
    }

    # Preserve joined choropleth data from existing GeoJSON
    if existing_props:
        for key in existing_props:
            if key not in props:
                props[key] = existing_props[key]

    # Add scoring results
    if score_result:
        props["composite_score"] = score_result["composite_score"]
        props["vigilent_classification"] = score_result["classification"]
        props["savings_per_mw"] = score_result["savings_per_mw"]
        props["payback_years"] = score_result["payback_years"]
        props["impact_on_opex_pct"] = score_result["impact_on_opex_pct"]
        props["estimated_savings"] = score_result["estimated_savings"]
        props["electricity_price_used"] = score_result["electricity_price_used"]
        props["operator_tier"] = score_result["operator_tier"]
        props["energy_pct_opex"] = score_result["energy_pct_opex"]
        props["capacity_factor"] = score_result["capacity_factor"]

    # Carry forward regional rate data for popups
    elec_key = REGION_ELEC_KEYS.get(region)
    water_key = REGION_WATER_KEYS.get(region)
    reg_key = REGION_REG_KEYS.get(region)
    if existing_props:
        if elec_key and elec_key in existing_props:
            props[elec_key] = existing_props[elec_key]
        if water_key and water_key in existing_props:
            props[water_key] = existing_props[water_key]
        if reg_key and reg_key in existing_props:
            props[reg_key] = existing_props[reg_key]

    return {
        "type": "Feature",
        "properties": props,
        "geometry": {"type": "Point", "coordinates": [lng, lat]}
    }


def write_geojson(region, features):
    """Write GeoJSON as qgis2web .js file."""
    fname, var_name = REGION_FILES[region]
    filepath = os.path.join(DATA_DIR, fname)
    geojson = {
        "type": "FeatureCollection",
        "name": var_name.replace("json_", ""),
        "features": features,
    }
    json_str = json.dumps(geojson, separators=(",", ":"))
    with open(filepath, "w") as f:
        f.write(f"var {var_name} = {json_str};")
    return filepath


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  BUILD MAP DATA — DYNAMIC PIPELINE")
    print("=" * 60)

    region_dcs = read_excel()

    total_scored = 0
    total_added = 0
    total_skipped = 0

    for region in ["US", "Canada", "Europe", "Brazil", "Other"]:
        dcs = region_dcs.get(region, [])
        if not dcs:
            continue

        print(f"\n  {region}: {len(dcs)} DCs in database")

        # Load existing GeoJSON for preserving joined data
        existing = load_existing_geojson(region)

        features = []
        scored = 0
        skipped = 0

        for dc in dcs:
            name = str(dc.get("Name", "")).strip()
            existing_props = existing.get(name)

            # Score the DC
            score_result = score_dc(dc, region)
            if score_result:
                scored += 1

            feature = build_feature(dc, region, score_result, existing_props)
            if feature is None:
                skipped += 1
                continue

            features.append(feature)

        filepath = write_geojson(region, features)
        size_kb = os.path.getsize(filepath) / 1024

        print(f"    Added: {len(features)}, Scored: {scored}, "
              f"Skipped (no coords): {skipped}")
        print(f"    File: {os.path.basename(filepath)} ({size_kb:.0f} KB)")

        total_scored += scored
        total_added += len(features)
        total_skipped += skipped

    print("\n" + "=" * 60)
    print(f"  Total: {total_added} DCs added, {total_scored} scored, "
          f"{total_skipped} skipped")
    print("  Done!")


if __name__ == "__main__":
    main()
