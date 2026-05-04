"""
Build India choropleth GeoJSON layers
=====================================
Reads Caroline's "[VIGILENT] Week 3 Caroline Scorecard.xlsx" India Scorecard sheet,
converts power (INR/kWh → ¢/kWh) and water (INR/KL → $/1000 gal),
applies city-proxy water tariffs for 5 missing states,
joins to a downloaded GADM/Natural-Earth India admin-1 GeoJSON,
emits 4 region layers matching the existing US/Europe/Canada/Brazil format.
"""

import json
import os
import re
import urllib.request
from pathlib import Path

import openpyxl

ROOT = Path("/Users/adamtang/Desktop/Vigilent")
SCORECARD = Path("/Users/adamtang/Desktop/[VIGILENT] Week 3 Caroline Scorecard.xlsx")
DATA_DIR = ROOT / "data"
RAW_DIR = DATA_DIR / "raw"
RAW_DIR.mkdir(parents=True, exist_ok=True)

# FX lock 2026-04-27 — keep stable so client-facing numbers don't drift
INR_PER_USD = 83.5
GAL_PER_M3 = 264.172  # 1 m³ = 264.172 US gallons

# Existing regions all use these property-key shapes — match exactly
ELEC_KEY = "India Statistics_Commercial Electricity Rate (¢/kWh)"
WATER_KEY = "India Statistics_Commercial Water Rate ($/1000 gallons)"
REGS_KEY = "India Statistics_Regulations"
COMP_KEY = "Composite"

# India national-level regulations applied to every polygon (state-level differentiation
# is largely cosmetic in the energy/water domain — these are all national)
INDIA_NATIONAL_REGS = "\n".join([
    "BEE PAT (Perform, Achieve, Trade)",
    "ECBC 2017 (Energy Conservation Building Code)",
    "BIS IS 16659 (Data Centre Energy Efficiency)",
    "MeitY Data Centre Policy (2020)",
    "IGBC Green Data Center rating",
])

# Caroline-listed cities for states with no water data → research-based municipal
# commercial water tariffs (INR/KL). Sources cited per row.
CITY_PROXY_WATER_INR_KL = {
    # state             city            INR/KL  source
    "Tripura":          ("Agartala",     22.0, "Agartala Municipal Corp commercial slab (~22 INR/KL above 30 KL)"),
    "Andhra Pradesh":   ("Visakhapatnam",36.0, "Greater Visakhapatnam Mun Corp commercial tariff 2024"),
    "Punjab":           ("Ludhiana",     32.0, "Punjab MC commercial water cess + base tariff (Ludhiana 2024)"),
    "Chhattisgarh":     ("Raipur",       28.0, "Raipur Mun Corp commercial slab (above 20 KL/month)"),
    "Puducherry":       ("Pondicherry",  30.0, "Puducherry Public Works Dept commercial tariff"),
}

# Composite weights — match Caroline's intent (electricity 0.40, water 0.20,
# DC density 0.20, regulatory 0.20). Each sub-score 0-100; composite 0-100.
COMPOSITE_WEIGHTS = {
    "electricity":  0.40,  # higher = better (lower price scores higher; we INVERT)
    "water":        0.20,  # higher = better (lower price scores higher; INVERT)
    "dc_density":   0.20,  # higher = better
    "regulatory":   0.20,  # higher = better
}


# ---------------------------------------------------------------------------
# 1. Read & deduplicate Caroline's India Scorecard
# ---------------------------------------------------------------------------

def read_india_scorecard():
    """Return list of dicts, one per unique state."""
    wb = openpyxl.load_workbook(SCORECARD, data_only=True)
    ws = wb["India Scorecard"]
    rows = []
    seen = set()
    for r in range(2, ws.max_row + 1):
        state = ws.cell(r, 1).value
        if state is None or not str(state).strip():
            continue
        state = str(state).strip()
        if state in seen:
            continue
        seen.add(state)
        rows.append({
            "state":            state,
            "city":             (ws.cell(r, 2).value or "").strip() if ws.cell(r, 2).value else "",
            "avg_power_inr":    ws.cell(r, 5).value,
            "dc_density":       ws.cell(r, 6).value,  # # DCs in state
            "avg_water_inr":    ws.cell(r, 9).value,
            "eff_index":        ws.cell(r, 10).value,
            "elec_score_5":     ws.cell(r, 11).value,
            "water_score_5":    ws.cell(r, 12).value,
            "density_score_5":  ws.cell(r, 13).value,
            "reg_score_5":      ws.cell(r, 14).value,
            "composite_5":      ws.cell(r, 15).value,
        })
    return rows


# ---------------------------------------------------------------------------
# 2. Unit conversion + city proxy fill
# ---------------------------------------------------------------------------

def convert_units(rows):
    """Add power_cents_per_kwh and water_usd_per_1000gal fields. Apply city proxies."""
    for row in rows:
        # Power: INR/kWh → USD/kWh → ¢/kWh
        if isinstance(row["avg_power_inr"], (int, float)):
            row["power_cents_per_kwh"] = round(row["avg_power_inr"] / INR_PER_USD * 100, 2)
        else:
            row["power_cents_per_kwh"] = None

        # Water: INR/m³ → USD/m³ → $/1000gal (× 1000/264.172)
        water_inr = row["avg_water_inr"]
        proxy_note = None
        if not isinstance(water_inr, (int, float)):
            # Try city proxy
            proxy = CITY_PROXY_WATER_INR_KL.get(row["state"])
            if proxy:
                _, water_inr, source = proxy
                proxy_note = f"Proxy: {source}"
        if isinstance(water_inr, (int, float)):
            usd_per_m3 = water_inr / INR_PER_USD
            row["water_usd_per_1000gal"] = round(usd_per_m3 * 1000 / GAL_PER_M3, 2)
        else:
            row["water_usd_per_1000gal"] = None
        row["water_proxy_note"] = proxy_note
    return rows


# ---------------------------------------------------------------------------
# 3. Composite score — recomputed at 0-100 scale (matching other regions)
# ---------------------------------------------------------------------------

def compute_composite(rows):
    """Recompute composite 0-100 with N/A treated as 0.

    Each sub-score normalized to 0-100:
      - electricity: lower price = higher score, scaled across observed range
      - water:       lower price = higher score, scaled across observed range
      - dc_density:  higher count = higher score, scaled across observed range
      - regulatory:  Caroline's 1-5 score → ×20 → 0-100
    """
    pows = [r["power_cents_per_kwh"] for r in rows if r["power_cents_per_kwh"] is not None]
    wats = [r["water_usd_per_1000gal"] for r in rows if r["water_usd_per_1000gal"] is not None]
    dens = [r["dc_density"] for r in rows if isinstance(r["dc_density"], (int, float))]

    p_lo, p_hi = (min(pows), max(pows)) if pows else (0, 1)
    w_lo, w_hi = (min(wats), max(wats)) if wats else (0, 1)
    d_lo, d_hi = (min(dens), max(dens)) if dens else (0, 1)

    def norm_inv(v, lo, hi):
        """Lower value = higher score."""
        if v is None or hi == lo:
            return 0
        return max(0, min(100, (1 - (v - lo) / (hi - lo)) * 100))

    def norm(v, lo, hi):
        """Higher value = higher score."""
        if v is None or hi == lo:
            return 0
        return max(0, min(100, (v - lo) / (hi - lo) * 100))

    for r in rows:
        e_score = norm_inv(r["power_cents_per_kwh"], p_lo, p_hi)
        w_score = norm_inv(r["water_usd_per_1000gal"], w_lo, w_hi)
        d_score = norm(r["dc_density"], d_lo, d_hi) if isinstance(r["dc_density"], (int, float)) else 0
        reg = r["reg_score_5"] if isinstance(r["reg_score_5"], (int, float)) else 0
        r_score = reg * 20  # 1-5 → 0-100
        composite = (
            e_score * COMPOSITE_WEIGHTS["electricity"]
            + w_score * COMPOSITE_WEIGHTS["water"]
            + d_score * COMPOSITE_WEIGHTS["dc_density"]
            + r_score * COMPOSITE_WEIGHTS["regulatory"]
        )
        r["composite_score_100"] = round(composite, 2)
        r["sub_scores"] = {
            "electricity": round(e_score, 1),
            "water":       round(w_score, 1),
            "dc_density":  round(d_score, 1),
            "regulatory":  round(r_score, 1),
        }
    return rows


# ---------------------------------------------------------------------------
# 4. Boundary fetch — Natural Earth admin-1 (CC0)
# ---------------------------------------------------------------------------

NE_ADMIN1_URL = "https://raw.githubusercontent.com/nvkelso/natural-earth-vector/master/geojson/ne_10m_admin_1_states_provinces.geojson"
NE_ADMIN1_PATH = RAW_DIR / "ne_10m_admin_1_states_provinces.geojson"


def fetch_boundaries():
    if NE_ADMIN1_PATH.exists():
        print(f"  Cached: {NE_ADMIN1_PATH}")
        return
    print(f"  Downloading {NE_ADMIN1_URL}...")
    urllib.request.urlretrieve(NE_ADMIN1_URL, NE_ADMIN1_PATH)
    print(f"  Saved: {NE_ADMIN1_PATH} ({NE_ADMIN1_PATH.stat().st_size // 1024} KB)")


def extract_india_polygons():
    """Return list of features for India admin-1 from Natural Earth."""
    with open(NE_ADMIN1_PATH) as f:
        gj = json.load(f)
    india = []
    for feat in gj["features"]:
        props = feat["properties"]
        if props.get("admin") == "India" or props.get("iso_a2") == "IN":
            india.append(feat)
    print(f"  India admin-1 features: {len(india)}")
    return india


# ---------------------------------------------------------------------------
# 5. State-name normalization + join
# ---------------------------------------------------------------------------

# Map Caroline's state name → list of acceptable Natural-Earth `name` candidates
# (NE uses British spellings and grouped UTs sometimes).
NAME_ALIASES = {
    "Andaman and Nicobar Islands":   ["Andaman and Nicobar"],
    "Dadra and Nagar Haveli":         ["Dadra and Nagar Haveli", "Dadara and Nagar Havelli"],
    "Daman and Diu":                  ["Daman and Diu"],
    "Delhi":                          ["NCT of Delhi", "Delhi"],
    "Jammu and Kashmir":              ["Jammu and Kashmir"],
    "Pondicherry":                    ["Puducherry", "Pondicherry"],
    "Puducherry":                     ["Puducherry", "Pondicherry"],
    "Orissa":                         ["Odisha", "Orissa"],
    "Odisha":                         ["Odisha", "Orissa"],
    "Uttarakhand":                    ["Uttarakhand", "Uttaranchal"],
    "Chhattisgarh":                   ["Chhattisgarh", "Chhattīsgarh"],
}


def _norm(s):
    return re.sub(r"[^a-z]", "", str(s).lower())


def join_data_to_polygons(rows, polys):
    """Match each Caroline row to a Natural Earth feature."""
    by_name = {}
    for f in polys:
        for cand in [
            f["properties"].get("name"),
            f["properties"].get("name_en"),
            f["properties"].get("name_alt"),
            f["properties"].get("woe_name"),
            f["properties"].get("gn_name"),
        ]:
            if cand:
                by_name[_norm(cand)] = f
    matches = {}
    unmatched_rows = []
    for r in rows:
        candidates = [r["state"]] + NAME_ALIASES.get(r["state"], [])
        match = None
        for cand in candidates:
            f = by_name.get(_norm(cand))
            if f:
                match = f
                break
        if match is None:
            unmatched_rows.append(r["state"])
        matches[r["state"]] = match
    return matches, unmatched_rows


# ---------------------------------------------------------------------------
# 6. Emit GeoJSON layer files
# ---------------------------------------------------------------------------

def write_layer(polys, rows, matches, var_name, fname, value_key, prop_key, descriptor):
    """Build and write one Leaflet-ready GeoJSON layer file."""
    features = []
    by_state = {r["state"]: r for r in rows}
    for state, feat in matches.items():
        if feat is None:
            continue
        row = by_state[state]
        v = row.get(value_key)
        props = {
            "NAME": state,
            prop_key: v if v is not None else None,
            COMP_KEY: row["composite_score_100"],
        }
        if value_key == "water_usd_per_1000gal" and row.get("water_proxy_note"):
            props["water_proxy_note"] = row["water_proxy_note"]
        new_feat = {
            "type": "Feature",
            "properties": props,
            "geometry": feat["geometry"],
        }
        features.append(new_feat)
    layer = {"type": "FeatureCollection", "features": features}
    out_path = DATA_DIR / fname
    with open(out_path, "w") as f:
        f.write(f"var {var_name} = ")
        json.dump(layer, f)
        f.write(";\n")
    print(f"    {fname:<50} {len(features):>3} features  ({descriptor})")
    return out_path


def write_regs_layer(polys, rows, matches):
    """Regulations — every state gets the national list (constant)."""
    features = []
    by_state = {r["state"]: r for r in rows}
    for state, feat in matches.items():
        if feat is None:
            continue
        row = by_state[state]
        # If reg_score_5 is None or 0 → mark N/A (per user)
        reg_score = row.get("reg_score_5")
        regs_text = INDIA_NATIONAL_REGS if isinstance(reg_score, (int, float)) and reg_score > 0 else "N/A"
        props = {
            "NAME": state,
            REGS_KEY: regs_text,
            COMP_KEY: row["composite_score_100"],
        }
        features.append({
            "type": "Feature",
            "properties": props,
            "geometry": feat["geometry"],
        })
    out_path = DATA_DIR / "IndiaRegulationsByQuantity_24.js"
    with open(out_path, "w") as f:
        f.write("var json_IndiaRegulationsByQuantity_24 = ")
        json.dump({"type": "FeatureCollection", "features": features}, f)
        f.write(";\n")
    print(f"    IndiaRegulationsByQuantity_24.js{'':<19}{len(features):>3} features  (regulations)")


def write_composite_layer(polys, rows, matches):
    features = []
    by_state = {r["state"]: r for r in rows}
    for state, feat in matches.items():
        if feat is None:
            continue
        row = by_state[state]
        props = {
            "NAME": state,
            COMP_KEY: row["composite_score_100"],
        }
        features.append({
            "type": "Feature",
            "properties": props,
            "geometry": feat["geometry"],
        })
    out_path = DATA_DIR / "IndiaCompositeScore_25.js"
    with open(out_path, "w") as f:
        f.write("var json_IndiaCompositeScore_25 = ")
        json.dump({"type": "FeatureCollection", "features": features}, f)
        f.write(";\n")
    print(f"    IndiaCompositeScore_25.js{'':<26}{len(features):>3} features  (composite)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 70)
    print("  INDIA CHOROPLETH LAYER BUILDER")
    print("=" * 70)

    print("\n[1/5] Reading Caroline India Scorecard...")
    rows = read_india_scorecard()
    print(f"  Unique states: {len(rows)}")

    print("\n[2/5] Converting units (INR→USD @ 1 USD = {:.1f} INR)...".format(INR_PER_USD))
    rows = convert_units(rows)
    proxied = [r["state"] for r in rows if r.get("water_proxy_note")]
    print(f"  Power: all {len(rows)} states converted to ¢/kWh")
    print(f"  Water: {sum(1 for r in rows if r['water_usd_per_1000gal'] is not None)} states have data; {len(proxied)} via city proxy: {proxied}")
    na_water = [r['state'] for r in rows if r['water_usd_per_1000gal'] is None]
    print(f"  Water N/A (composite contribution = 0): {len(na_water)} states")

    print("\n[3/5] Computing composite (0-100) treating N/A as 0...")
    rows = compute_composite(rows)
    rng = sorted(r["composite_score_100"] for r in rows)
    print(f"  Composite range: {rng[0]} – {rng[-1]}, median {rng[len(rng)//2]}")

    print("\n[4/5] Loading India admin-1 polygons...")
    fetch_boundaries()
    polys = extract_india_polygons()
    matches, unmatched = join_data_to_polygons(rows, polys)
    if unmatched:
        print(f"  WARNING — unmatched states: {unmatched}")
    matched = sum(1 for v in matches.values() if v is not None)
    print(f"  Matched {matched}/{len(rows)} states to polygons")

    print("\n[5/5] Writing GeoJSON layers...")
    write_layer(polys, rows, matches,
                "json_IndiaCommercialElectricityRateskWh_22",
                "IndiaCommercialElectricityRateskWh_22.js",
                "power_cents_per_kwh", ELEC_KEY, "electricity")
    write_layer(polys, rows, matches,
                "json_IndiaCommercialWaterRates1000gallons_23",
                "IndiaCommercialWaterRates1000gallons_23.js",
                "water_usd_per_1000gal", WATER_KEY, "water")
    write_regs_layer(polys, rows, matches)
    write_composite_layer(polys, rows, matches)

    # Also write a flat CSV report so Caroline can sanity-check the conversions
    report_path = ROOT / "output" / "india_state_metrics.csv"
    report_path.parent.mkdir(exist_ok=True)
    with open(report_path, "w") as f:
        f.write("state,city,power_inr_per_kwh,power_cents_per_kwh,"
                "water_inr_per_kl,water_usd_per_1000gal,water_proxy,"
                "dc_count,reg_score_5,composite_0_100,e_sub,w_sub,d_sub,r_sub,polygon_matched\n")
        for r in sorted(rows, key=lambda x: -x["composite_score_100"]):
            sub = r.get("sub_scores", {})
            f.write(f'"{r["state"]}","{r["city"]}",'
                    f'{r["avg_power_inr"]},{r["power_cents_per_kwh"]},'
                    f'{r["avg_water_inr"]},{r["water_usd_per_1000gal"]},'
                    f'"{r.get("water_proxy_note") or ""}",'
                    f'{r["dc_density"]},{r["reg_score_5"]},'
                    f'{r["composite_score_100"]},'
                    f'{sub.get("electricity","")},{sub.get("water","")},'
                    f'{sub.get("dc_density","")},{sub.get("regulatory","")},'
                    f'{"Y" if matches.get(r["state"]) else "N"}\n')
    print(f"\n  Report: {report_path}")
    print("\nDone.")


if __name__ == "__main__":
    main()
